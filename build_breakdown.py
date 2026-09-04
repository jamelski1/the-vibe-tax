"""Build The_Vibe_Tax_Problem_Breakdown.xlsx — per-problem/per-model success/fail
across HumanEval, HumanEval+, and LiveCodeBench, with the model's actual code."""
import ast, json, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os as _os; ROOT = _os.path.dirname(_os.path.abspath(__file__))
OUT = os.path.join(ROOT, "The_Vibe_Tax_Problem_Breakdown.xlsx")
REP = "agentic_terse"            # representative condition (framing is null)
CAP = 3000                       # max chars for long text cells

# ---------- topic classifier (same as analyze_lcb_capability.py) ----------
TOPICS = [
    ("dynamic programming", r"\b(dynamic programming|dp\b|subsequence|number of ways|minimum cost|maximum (sum|score|value)|partitions?)\b"),
    ("graph", r"\b(graph|node|edge|adjacen|connected component|shortest path|cities|roads?|network)\b"),
    ("tree", r"\b(binary tree|tree|root|leaf|subtree|ancestor|parent node)\b"),
    ("string", r"\b(string|substring|palindrome|character|prefix|suffix|anagram|lexicograph)\b"),
    ("intervals/sorting", r"\b(interval|sort|sorted|merge|schedule|meeting|overlap)\b"),
    ("greedy/array", r"\b(array|subarray|greedy|adjacent|window|two pointers?)\b"),
    ("math/number", r"\b(prime|gcd|lcm|modulo|divisor|digits?|binary representation|factorial|arithmetic)\b"),
    ("bit manipulation", r"\b(bit|xor|bitwise|set bits)\b"),
    ("simulation/geometry", r"\b(simulat|grid|matrix|coordinate|move|direction|snake|robot|game)\b"),
]
def classify(t):
    t = t.lower()
    for name, pat in TOPICS:
        if re.search(pat, t): return name
    return "other"

def lcb_code(completion, entry):
    if not completion: return ""
    cands = []
    for b in re.findall(r"```(?:python|py)?\s*\n(.*?)```", completion, re.DOTALL):
        if "class Solution" in b or f"def {entry}" in b: cands.append(b)
    text = "\n".join(l for l in completion.split("\n") if not l.strip().startswith("```"))
    for anchor in ("class Solution", f"def {entry}"):
        i = text.find(anchor)
        if i != -1: cands.append(text[i:])
    for c in cands:
        lines = c.split("\n")
        while lines:
            src = "\n".join(lines)
            try:
                ast.parse(src)
                if "class Solution" in src or f"def {entry}" in src: return src
                break
            except SyntaxError: lines.pop()
    return (completion or "")[:CAP]

def clip(s):
    s = s or ""
    return s if len(s) <= CAP else s[:CAP] + "\n…(truncated)"

def load(p): return json.load(open(os.path.join(ROOT, p), encoding="utf-8"))

# ---------- styling ----------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Arial", color="006100", bold=True)
FAIL_FONT = Font(name="Arial", color="9C0006", bold=True)
BASE_FONT = Font(name="Arial", size=10)
MONO = Font(name="Consolas", size=9)
THIN = Border(*[Side(style="thin", color="D9D9D9")]*4)
TOP = Alignment(vertical="top", wrap_text=True)

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c); cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    ws.row_dimensions[1].height = 22

def write_rows(ws, headers, rows, widths, result_cols=()):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # body font + borders + wrap
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BASE_FONT; cell.alignment = TOP; cell.border = THIN
    # code columns monospace (last col assumed code if wide)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if ws.column_dimensions[cell.column_letter].width >= 60:
                cell.font = MONO; cell.alignment = TOP
    # color PASS/FAIL cells
    for ci in result_cols:
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                v = str(cell.value)
                if v == "PASS": cell.fill = PASS_FILL; cell.font = PASS_FONT
                elif v == "FAIL": cell.fill = FAIL_FILL; cell.font = FAIL_FONT
    style_header(ws, len(headers))

wb = openpyxl.Workbook()

# ============ README ============
ws = wb.active; ws.title = "README"
readme = [
 ("The Vibe Tax — Problem-by-Problem Breakdown", 16, True),
 ("", 10, False),
 ("Per-problem, per-model success/failure across three benchmarks. One tab per benchmark.", 11, False),
 ("Framing is null (terse/casual/polite/multilingual do not differ), so each row shows the", 11, False),
 ("TERSE condition as representative; the LCB tabs also give a pass-count across all 4 framings.", 11, False),
 ("", 10, False),
 ("Tabs", 13, True),
 ("• LCB — one row per (problem × model). difficulty, topic, the problem, PASS/FAIL, the model's code.", 11, False),
 ("• LCB by problem — one row per problem: how many of the 12 attempts (4 framings × 3 models) solved it.", 11, False),
 ("• HumanEval — base HumanEval tests (saturated; ~all pass). Model code shown.", 11, False),
 ("• HumanEval+ — same problems, EvalPlus edge-case tests (harder). base vs plus pass shown.", 11, False),
 ("", 10, False),
 ("Key columns", 13, True),
 ("• result / PASS-FAIL: did the model's code pass the graded tests for that problem.", 11, False),
 ("• model_solution: the actual code the model produced (robustly extracted; this is what was run).", 11, False),
 ("• topic (LCB): keyword-derived, directional not gold-standard (LCB ships no topic labels).", 11, False),
 ("", 10, False),
 ("What to look for", 13, True),
 ("• LCB failures are 100% wrong-logic: valid code, wrong answer — the model can code, not solve.", 11, False),
 ("• Difficulty drives success: easy ~99%, medium ~81%, hard ~51%. Dynamic programming is the weakest topic.", 11, False),
 ("• HumanEval is saturated (~97%) — few failures to see there; LCB is where the signal is.", 11, False),
 ("", 10, False),
 ("Sources: data/vibe_tax_lcb/lcb_scored.json + lcb_v3_responses.json + lcb_problems.jsonl;", 9, False),
 ("data/vibe_tax_v3/vibe_tax_v3_scored.json + _plus_scored.json + _responses.json. Numbers are pass@1, temp 0.", 9, False),
]
for i,(txt,sz,bold) in enumerate(readme,1):
    c = ws.cell(i,1,txt); c.font = Font(name="Arial", size=sz, bold=bold)
ws.column_dimensions["A"].width = 110

# ============ LCB ============
lcb_scored = load("data/vibe_tax_lcb/lcb_scored.json")
lcb_resp = {(x["task_id"],x["level"],x["model"]): x for x in load("data/vibe_tax_lcb/lcb_v3_responses.json")}
probs = {json.loads(l)["task_id"]: json.loads(l) for l in open(os.path.join(ROOT,"data/vibe_tax_lcb/lcb_problems.jsonl"))}
for p in probs.values(): p["topic"] = classify(p["question_content"]+" "+p["entry_point"])

# passed lookup and per-(problem,model) count across framings
passed = {(x["task_id"],x["level"],x["model"]): x["passed"] for x in lcb_scored}
from collections import defaultdict
cnt = defaultdict(lambda:[0,0])
for x in lcb_scored:
    cnt[(x["task_id"],x["model"])][0]+=x["passed"]; cnt[(x["task_id"],x["model"])][1]+=1

rows=[]
for tid in sorted(probs):
    p=probs[tid]
    for model in ("chatgpt","claude","codestral"):
        pv = passed.get((tid,REP,model))
        if pv is None: continue
        k,n = cnt[(tid,model)]
        comp = lcb_resp.get((tid,REP,model),{}).get("completion")
        rows.append([tid, p["difficulty"], p["topic"], p["entry_point"], model,
                     "PASS" if pv else "FAIL", f"{k}/{n}",
                     clip(p["question_content"]), lcb_code(comp,p["entry_point"])])
# sort: hardest + most-failed first for quick insight
order={"hard":0,"medium":1,"easy":2}
rows.sort(key=lambda r:(order.get(r[1],3), r[5]!="FAIL", r[0], r[4]))
ws = wb.create_sheet("LCB")
write_rows(ws,
    ["task_id","difficulty","topic","method","model","result (terse)","passed /4 (all framings)","problem_statement","model_solution"],
    rows, [11,9,18,26,10,13,20,70,80], result_cols=(6,))

# ============ LCB by problem ============
rows=[]
for tid in sorted(probs):
    p=probs[tid]
    per={m:cnt.get((tid,m),[0,0]) for m in ("chatgpt","claude","codestral")}
    tot=sum(v[0] for v in per.values()); den=sum(v[1] for v in per.values())
    verdict = "Always" if tot==den and den>0 else ("Never" if tot==0 else "Partial")
    rows.append([tid,p["difficulty"],p["topic"],p["entry_point"],
                 f'{per["chatgpt"][0]}/{per["chatgpt"][1]}',
                 f'{per["claude"][0]}/{per["claude"][1]}',
                 f'{per["codestral"][0]}/{per["codestral"][1]}',
                 f"{tot}/{den}", verdict])
rows.sort(key=lambda r:(order.get(r[1],3), int(r[7].split("/")[0])))
ws = wb.create_sheet("LCB by problem")
write_rows(ws,
    ["task_id","difficulty","topic","method","chatgpt /4","claude /4","codestral /4","total /12","verdict"],
    rows, [11,9,18,30,11,11,12,10,10])
# color verdict
for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
    for cell in row:
        if cell.value=="Always": cell.fill=PASS_FILL; cell.font=PASS_FONT
        elif cell.value=="Never": cell.fill=FAIL_FILL; cell.font=FAIL_FONT

# ============ HumanEval + HumanEval+ ============
he_probs={}
for l in open(os.path.join(ROOT,"data/HumanEval.jsonl/human-eval-v2-20210705.jsonl")):
    d=json.loads(l); he_probs[d["task_id"]]=d
v3_resp={(x["task_id"],x["level"],x["model"]):x for x in load("data/vibe_tax_v3/vibe_tax_v3_responses.json")}
v3_base={(x["task_id"],x["level"],x["model"]):x["passed"] for x in load("data/vibe_tax_v3/vibe_tax_v3_scored.json")}
v3_plus={(x["task_id"],x["level"],x["model"]):x["passed"] for x in load("data/vibe_tax_v3/vibe_tax_v3_plus_scored.json")}

def he_rows(passed_map, second_map=None):
    rows=[]
    keys=sorted({(t,m) for (t,l,m) in v3_resp if l==REP})
    for tid,model in keys:
        pv=passed_map.get((tid,REP,model))
        if pv is None: continue
        comp=v3_resp.get((tid,REP,model),{}).get("completion")
        ep=he_probs.get(tid,{}).get("entry_point","")
        base=[tid,ep,model,"PASS" if pv else "FAIL"]
        if second_map is not None:
            sv=second_map.get((tid,REP,model))
            base.append("PASS" if sv else "FAIL")
        base += [clip(he_probs.get(tid,{}).get("prompt","")), clip(comp)]
        rows.append(base)
    rows.sort(key=lambda r:(r[3]!="FAIL", r[0], r[2]))
    return rows

ws=wb.create_sheet("HumanEval")
write_rows(ws, ["task_id","entry_point","model","result (terse)","problem (spec)","model_solution"],
           he_rows(v3_base), [12,26,10,13,70,80], result_cols=(4,))

ws=wb.create_sheet("HumanEval+")
rows=he_rows(v3_plus, second_map=None)
# rebuild with base vs plus both
rows=[]
for tid,model in sorted({(t,m) for (t,l,m) in v3_resp if l==REP}):
    b=v3_base.get((tid,REP,model)); pl=v3_plus.get((tid,REP,model))
    if pl is None: continue
    comp=v3_resp.get((tid,REP,model),{}).get("completion"); ep=he_probs.get(tid,{}).get("entry_point","")
    rows.append([tid,ep,model,"PASS" if b else "FAIL","PASS" if pl else "FAIL",
                 clip(he_probs.get(tid,{}).get("prompt","")),clip(comp)])
rows.sort(key=lambda r:(r[4]!="FAIL", r[0], r[2]))
write_rows(ws, ["task_id","entry_point","model","base result","plus result (edge tests)","problem (spec)","model_solution"],
           rows, [12,26,10,12,16,70,80], result_cols=(4,5))

wb.save(OUT)
print("wrote", OUT)
for s in wb.sheetnames: print("  tab:", s, "rows:", wb[s].max_row-1 if s!="README" else "-")
